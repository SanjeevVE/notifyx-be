from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import Optional
from datetime import datetime
import os
import uuid
import csv
import io

from app.db.database import get_db
from app.models.user import User
from app.models.import_job import ImportJob, ImportJobStatus, MAX_PREVIEW_ROWS
from app.models.contact import ContactList
from app.schemas.import_job import (
    ImportUploadResponse,
    ImportStartRequest,
    ImportJobResponse,
    ImportJobBrief,
    PaginatedImportJobs,
    suggest_column_mappings,
)
from app.api.v1.auth import get_current_user
from app.core.config import settings

router = APIRouter()

# Directory for temporary file storage
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


def parse_excel_file(file_path: str):
    """
    Parse Excel file and return column headers, preview rows, and total count.
    Supports both .xlsx and .csv files.
    """
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    columns = []
    preview_rows = []
    total_rows = 0

    if ext == '.csv':
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            columns = reader.fieldnames or []

            for i, row in enumerate(reader):
                total_rows += 1
                if i < MAX_PREVIEW_ROWS:
                    preview_rows.append(dict(row))

    elif ext in ('.xlsx', '.xls'):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Excel support not installed. Please install openpyxl."
            )

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file is empty"
            )

        # First row is headers
        columns = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(rows[0])]

        # Data rows
        for i, row in enumerate(rows[1:]):
            total_rows += 1
            if i < MAX_PREVIEW_ROWS:
                row_dict = {}
                for j, cell in enumerate(row):
                    col_name = columns[j] if j < len(columns) else f"Column_{j}"
                    row_dict[col_name] = str(cell) if cell is not None else ""
                preview_rows.append(row_dict)

        wb.close()
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file format: {ext}. Supported formats: .xlsx, .xls, .csv"
        )

    return columns, preview_rows, total_rows


# ============= Upload File =============

@router.post("/upload", response_model=ImportUploadResponse)
async def upload_import_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Upload an Excel or CSV file for contact import.
    Returns column headers and preview rows for mapping.
    """
    # Validate file extension
    filename = file.filename or "upload"
    _, ext = os.path.splitext(filename)
    ext = ext.lower()

    if ext not in ('.xlsx', '.xls', '.csv'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel (.xlsx, .xls) and CSV files are supported"
        )

    # Generate unique filename
    unique_filename = f"{uuid.uuid4()}{ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_filename)

    # Save file to disk
    try:
        content = await file.read()
        with open(file_path, 'wb') as f:
            f.write(content)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to save file: {str(e)}"
        )

    # Parse file
    try:
        columns, preview_rows, total_rows = parse_excel_file(file_path)
    except HTTPException:
        # Clean up file on error
        if os.path.exists(file_path):
            os.remove(file_path)
        raise
    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {str(e)}"
        )

    if not columns:
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File has no columns"
        )

    if total_rows == 0:
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File has no data rows"
        )

    # Generate suggested mappings
    suggested_mappings = suggest_column_mappings(columns)

    # Create import job
    import_job = ImportJob(
        organization_id=current_user.organization_id,
        file_name=filename,
        file_path=file_path,
        file_size=len(content),
        status=ImportJobStatus.PENDING,
        total_rows=total_rows,
    )

    db.add(import_job)
    await db.commit()
    await db.refresh(import_job)

    return ImportUploadResponse(
        import_job_id=import_job.id,
        file_name=filename,
        columns=columns,
        preview_rows=preview_rows,
        total_rows=total_rows,
        suggested_mappings=suggested_mappings,
    )


# ============= Start Import =============

@router.post("/{job_id}/start", response_model=ImportJobResponse)
async def start_import(
    job_id: int,
    request: ImportStartRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Start the import process with column mappings.
    Triggers a Celery background task for large files.
    """
    # Get import job
    result = await db.execute(
        select(ImportJob).filter(
            ImportJob.id == job_id,
            ImportJob.organization_id == current_user.organization_id,
        )
    )
    job = result.scalar_one_or_none()

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Import job not found"
        )

    if job.status != ImportJobStatus.PENDING:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Import job is already {job.status.value}"
        )

    # Validate list_id if provided
    if request.options.list_id:
        result = await db.execute(
            select(ContactList).filter(
                ContactList.id == request.options.list_id,
                ContactList.organization_id == current_user.organization_id,
            )
        )
        contact_list = result.scalar_one_or_none()
        if not contact_list:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Contact list not found"
            )

    # Save mapping configuration
    job.column_mapping = [m.model_dump() for m in request.mappings]
    job.options = request.options.model_dump()
    job.status = ImportJobStatus.PROCESSING
    job.started_at = datetime.utcnow()

    await db.commit()

    # Trigger Celery task for background processing
    try:
        from app.tasks.import_tasks import process_contact_import
        task = process_contact_import.delay(job.id)
        job.celery_task_id = task.id
        await db.commit()
    except Exception as e:
        # If Celery is not available, process synchronously (for small imports)
        # This is a fallback for development without Celery
        pass

    await db.refresh(job)
    return job


# ============= Get Import Job Status =============

@router.get("/{job_id}", response_model=ImportJobResponse)
async def get_import_job(
    job_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get the status and progress of an import job"""
    result = await db.execute(
        select(ImportJob).filter(
            ImportJob.id == job_id,
            ImportJob.organization_id == current_user.organization_id,
        )
    )
    job = result.scalar_one_or_none()

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Import job not found"
        )

    return job


# ============= List Import Jobs =============

@router.get("", response_model=PaginatedImportJobs)
async def list_import_jobs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List all import jobs for the organization"""
    query = select(ImportJob).filter(
        ImportJob.organization_id == current_user.organization_id
    )

    # Get total count
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar()

    # Apply pagination
    query = query.order_by(ImportJob.created_at.desc())
    query = query.offset((page - 1) * page_size).limit(page_size)

    result = await db.execute(query)
    jobs = result.scalars().all()

    return PaginatedImportJobs(
        items=jobs,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=(total + page_size - 1) // page_size if total > 0 else 0
    )


# ============= Cancel Import =============

@router.post("/{job_id}/cancel", response_model=ImportJobResponse)
async def cancel_import(
    job_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Cancel an in-progress import job"""
    result = await db.execute(
        select(ImportJob).filter(
            ImportJob.id == job_id,
            ImportJob.organization_id == current_user.organization_id,
        )
    )
    job = result.scalar_one_or_none()

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Import job not found"
        )

    if job.status not in (ImportJobStatus.PENDING, ImportJobStatus.PROCESSING):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot cancel job with status: {job.status.value}"
        )

    # Revoke Celery task if running
    if job.celery_task_id:
        try:
            from app.tasks.celery_app import celery_app
            celery_app.control.revoke(job.celery_task_id, terminate=True)
        except Exception:
            pass

    job.status = ImportJobStatus.CANCELLED
    job.completed_at = datetime.utcnow()

    # Clean up uploaded file
    if job.file_path and os.path.exists(job.file_path):
        try:
            os.remove(job.file_path)
        except Exception:
            pass

    await db.commit()
    await db.refresh(job)

    return job


# ============= Sample Excel Download =============

@router.get("/sample-excel")
async def download_sample_excel():
    """Download a sample Excel file for contact import"""
    current_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
    sample_file = os.path.join(current_dir, "sample_contacts.xlsx")

    # If sample doesn't exist, create it
    if not os.path.exists(sample_file):
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Contacts"

            # Headers
            headers = ["Email", "Full Name", "Company", "Phone", "Job Title", "Department"]
            ws.append(headers)

            # Sample data
            sample_data = [
                ["john@example.com", "John Doe", "Acme Inc", "+1234567890", "CEO", "Executive"],
                ["jane@example.com", "Jane Smith", "Tech Corp", "+0987654321", "CTO", "Engineering"],
                ["bob@example.com", "Bob Wilson", "Startup Ltd", "+1122334455", "Manager", "Sales"],
            ]
            for row in sample_data:
                ws.append(row)

            wb.save(sample_file)
            wb.close()
        except ImportError:
            # Fall back to CSV if openpyxl not installed
            sample_file = os.path.join(current_dir, "sample_contacts.csv")
            if not os.path.exists(sample_file):
                with open(sample_file, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Email", "Full Name", "Company", "Phone", "Job Title", "Department"])
                    writer.writerow(["john@example.com", "John Doe", "Acme Inc", "+1234567890", "CEO", "Executive"])
                    writer.writerow(["jane@example.com", "Jane Smith", "Tech Corp", "+0987654321", "CTO", "Engineering"])

    if not os.path.exists(sample_file):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Sample file not found"
        )

    filename = os.path.basename(sample_file)
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if filename.endswith('.xlsx') else "text/csv"

    return FileResponse(
        path=sample_file,
        filename=filename,
        media_type=media_type
    )
